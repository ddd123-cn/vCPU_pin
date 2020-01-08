# -*- coding: utf-8 -*-
"""
Created on Sat Jan  4 11:38:57 2020

@author: ejiandi
"""
#from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.comments import Comment
wb = Workbook()
ws = wb.active
ft_bold = Font(b=True,size=14)
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="0000ff")

ws['A1'] = "Compute"
ws['A2'] = "CPU"
ws['A1'].font = ft_bold
ws['A2'].font = ft_bold
#读取txt文件中的vcpu分布信息
host_list = []
VNF_list = []
#vm_list = []
cell_pos_list = []
f = open ('vcpu_pin.txt','r')
for line in f.readlines():
    if line != " \n":
        if line[0:7] == "compute":
            print(line)
            host_list.append(line.rstrip('\n'))
            column1 = len(host_list) * 2
        else:
            print(line)
            vm = line.split()[1]
            #vm_list.append(vm)
            vnf = vm.split("_")[0]
            vnf = vnf.split("-")[0]
            #vnf = vnf.split(">")[-1]
            VNF_list.append(vnf)
            cpulist = line.split("'")[1]
            cpulist = cpulist.split(",")
            cpuset = []
            cell_set = []
            for id in range(len(cpulist)//2):
                cpuset.append([cpulist[0 + id],cpulist[len(cpulist)//2 +id]])
            for cpupair in cpuset:
                cpupair = [int(cpupair[0]),int(cpupair[1])]
                if (cpupair[0] % 2) == 0:
                    row1 = 3 + cpupair[0]//2
                else:
                    row1 = 2 + cpupair[1]//2
                cell_pos = [row1,column1]
                cell_set.append(cell_pos)
            cell_pos_list.append(cell_set)
f.close()
#画表格
cpu_core = cpupair[1] - cpupair[0]
for id in range(len(host_list)):
    ws.merge_cells(start_row=1,start_column=id*2+2,end_row=1,end_column=id*2+3)
    ws.cell(row=1,column=id*2+2).value = host_list[id]
    ws.cell(row=1,column=id*2+2).border = Border(top=double, left=double, right=thin, bottom=double)
    ws.cell(row=1,column=id*2+3).border = Border(top=double, left=double, right=thin, bottom=double)
    ws.cell(row=1,column=id*2+2).fill = PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=1,column=id*2+2).font  = Font(b=True, color="FF0000")
    ws.cell(row=1,column=id*2+2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=3,column=id*2+2).fill = PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=3,column=id*2+3).fill = PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=3,column=id*2+3).border = Border(top=double)
    ws.cell(row=3+cpu_core//2,column=id*2+2).fill = PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=3+cpu_core//2,column=id*2+3).fill = PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=3+cpu_core//2,column=id*2+3).border = Border(top=double)

    for row_ in range(cpu_core//2):
        ws.cell(row=3+row_,column=id*2+2).value = row_ *2
        ws.cell(row=3+row_,column=id*2+2).border = Border(left=double)
        ws.cell(row=3+row_,column=id*2+3).value = row_ *2 + cpu_core
        ws.cell(row=3+row_+cpu_core//2,column=id*2+2).value = row_ *2 + 1
        ws.cell(row=3+row_+cpu_core//2,column=id*2+2).border = Border(left=double)
        ws.cell(row=3+row_+cpu_core//2,column=id*2+3).value = row_ *2 + cpu_core + 1
    ws.cell(row=3,column=id*2+2).border = Border(top=double,left=double)
    ws.cell(row=3+cpu_core//2,column=id*2+2).border = Border(top=double,left=double)

ws['A3'] = "NUMA_0"
ws.merge_cells(start_row=3,start_column=1,end_row=(2+cpu_core//2),end_column=1)
ws.merge_cells(start_row=3+cpu_core//2,start_column=1,end_row=2+cpu_core,end_column=1)
ws.cell(row=3+cpu_core//2,column=1).value = "NUMA_1"
ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=3+cpu_core//2,column=1).alignment = Alignment(horizontal="center", vertical="center")
ws['A3'].border = Border(top=double)
ws.cell(row=3+cpu_core//2,column=1).border = Border(top=double)
#上色
colormap = {"default":0}
VNFs = list(set(VNF_list))
VNFs.sort()
###图例
for id in range(len(VNFs)):
    colormap[VNFs[id]] = 42 + id
    color_ = colormap[VNFs[id]]
    Fill_ = PatternFill("solid",fgColor=Color(indexed=color_))
#    ws.cell(row=4+id, column=(len(host_list)*2+3)).value = VNFs[id]+'  '+str(color_)
    ws.cell(row=4+id, column=(len(host_list)*2+3)).value = VNFs[id]
    ws.cell(row=4+id, column=(len(host_list)*2+3)).fill = Fill_
#涂格子
for id in range(len(VNF_list)):
    color_ = colormap[VNF_list[id]]
    Fill_ = PatternFill("solid",fgColor=Color(indexed=color_))
    comment = Comment(VNF_list[id],'ejiandi')
    cellset = cell_pos_list[id]
    for id in cellset:
        ws.cell(row=id[0],column=id[1]).fill = Fill_
        ws.cell(row=id[0],column=id[1]+1).fill = Fill_
        ws.cell(row=id[0],column=id[1]+1).comment = comment
wb.save('testvcpu_pin.xlsx')